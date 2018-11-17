# -*- coding: utf-8 -*-
import os
import glob
import re
import copy
import map
import openpyxl
from pprint import pprint

# compare and give similarity score
def how_much_similar(base, target):
    score = 0
    weightL = 10
    weightW = 10
    invalidKey = ['Travel', 'sky']
    for baseL in base['label']:
        for targetL in target['label']:
            # len(baseL['description']) != 0 and len(targetL['description']) != 0
            if len(baseL['description']) != 0 and len(targetL['description']) != 0 \
            and targetL['description'] == baseL['description'] \
            and baseL['description'] not in invalidKey:
                score = score + weightL * (baseL['score'] + targetL['score'])
    for baseW in base['result-web']:
        for targetW in target['result-web']:
            # len(baseW['description']) != 0 and len(targetW['description']) != 0
            if len(baseW['description']) != 0 and len(targetW['description']) != 0 \
            and targetW['description'] == baseW['description'] \
            and baseW['description'] not in invalidKey:
                score = score + weightW * (baseW['score'] + targetW['score'])
    return score

# consider the distance between user-input-place and 80%-candidates
def sort_with_distance(target_list, _user_lat, _user_lon):
    # # for check
    # print('\nsort_with_distance start, target_list: ')
    # pprint(target_list)
    # reconstitute target_list: by making groups or just sorting target_list
    # according to len(target_list)

    wb = openpyxl.load_workbook('placeName.xlsx')
    ws = wb.active

    if len(target_list) > 2:
        # calculate avg. of similarity-difference
        sum = 0
        for i in range(0, len(target_list)-1):
            sum = sum + (target_list[i]['score'] - target_list[i+1]['score'])
        avg = sum / (len(target_list)-1)
        # make group
        group_list = []
        temp_group = []
        for target in target_list:
            if target_list.index(target) == 0:
                temp_group.append(target)
            elif temp_group[-1]['score'] - target['score'] <= avg:
                temp_group.append(target)
            else:
                group_list.append(temp_group)
                temp_group = []
                temp_group.append(target)
        # # for check
        # print('group_list: ')
        # pprint(group_list)
        # get distance and sort with distance in the same group
        for group in group_list:
            for target in group:
                # add target['fileNum'] & target['linkFinder']
                p = re.compile('[^0-9]')
                temp = os.path.splitext(target['fileName'])
                temp = os.path.split(temp[0])
                target['fileNum'] = filter(lambda x: x.isdigit(), temp[1]) # number in fileName
                target['linkFinder'] = ''.join(p.findall(temp[1])) # place name in fileName
                # save target['placeName'] using excel file
                target['placeName'] = 'null' # default value
                for i in range(1, ws.max_row + 1):
                    if target['linkFinder'].lower() == ws.cell(row=i, column=1).value.lower():
                        target['placeName'] = ws.cell(row=i, column=2).value
                        break;
                # using target['linkFinder'], get lat & lon of target
                # targetInfo = map.get_where(target['linkFinder']) # targetInfo = {'name': '', 'address':'', 'lat':'', 'lon':''}
                if target['placeName'] != 'null':
                    targetInfo = map.get_where(target['placeName'])
                else:
                    targetInfo = map.get_where(target['linkFinder'])
                    target['placeName'] = targetInfo['name']
                # save targetInfo
                target['placeAddress'] = targetInfo['address']
                # entering each lat & lon of target and user-place, get the distance betw. two places
                target['distance'] = map.get_D(_user_lat, _user_lon, targetInfo['lat'], targetInfo['lon'])
            # sort targets with ascending distance within group
            group = sorted(group, key=lambda d: (d['distance']), reverse=False)
        # update target_list
        target_list = []
        for group in group_list:
            for target in group:
                target_list.append(target)
    else:
        for target in target_list:
            # add target['fileNum'] & target['linkFinder']
            p = re.compile('[^0-9]')
            temp = os.path.splitext(target['fileName'])
            temp = os.path.split(temp[0])
            target['fileNum'] = filter(lambda x: x.isdigit(), temp[1]) # number in fileName
            target['linkFinder'] = ''.join(p.findall(temp[1])) # place name in fileName
            # using target['linkFinder'], get lat & lon of target
            # targetInfo = map.get_where(target['linkFinder'])
            # save target['placeName'] using excel file
            target['placeName'] = 'null' # default value
            for i in range(1, ws.max_row + 1):
                if target['linkFinder'].lower() == ws.cell(row=i, column=1).value.lower():
                    target['placeName'] = ws.cell(row=i, column=2).value
                    break;
            if target['placeName'] != 'null':
                    targetInfo = map.get_where(target['placeName'])
            else:
                targetInfo = map.get_where(target['linkFinder'])
                target['placeName'] = targetInfo['name']
            # save targetInfo
            # target['placeName'] = targetInfo['name']
            target['placeAddress'] = targetInfo['address']
            # entering each lat & lon of target and user-place, get the distance betw. two places
            target['distance'] = map.get_D(_user_lat, _user_lon, targetInfo['lat'], targetInfo['lon'])
        print('\nlength of target_list <= 2, just sort')
        target_list = sorted(target_list, key=lambda d: (d['distance']), reverse=False)
    # # for check
    # print('sort_with_distance end, target_list: ')
    # pprint(target_list)
    return target_list

def print_result(target_list, _APP_ROOT):
    best = []
    overlap = []
    # # for check
    # print 'print_result start: '
    # pprint(target_list)
    print '\n'
    # print result
    for target in target_list:
        # add to best only when len(best) < 10 and (at first time or target['fileName'] not in overlap)
        if len(best) < 9 and (target_list.index(target) == 0 or all(o not in target['fileName'] for o in overlap)):
            # add target to best and linkFinder to overlap
            best.append(target)
            overlap.append(target['linkFinder'])
    # load links
    bestLink_list = [] # [ [links for place 1], [links for place 2], [links for place 3] ]
    link_target = os.path.join(_APP_ROOT, 'URL/') # directory's address which in link fils exist
    for target in best:
        linkFiles = os.path.join(link_target, target['linkFinder']) # directory-path in which link files of single place are
        linkDirList = glob.glob(linkFiles + '*.txt') # full-path for link files of single place
        # remove first link (later added at first)
        firstLinkDir = link_target + target['linkFinder'] + target['fileNum'] + '.txt'
        linkDirList.remove(firstLinkDir)
        # load every links
        temp = []
        for linkDir in linkDirList:
            f = open(linkDir, 'r')
            link = f.read().splitlines()
            temp.append(link[0])
        # load first link and add at first
        f = open(firstLinkDir, 'r')
        link = f.read().splitlines()
        temp.insert(0, link[0])
        bestLink_list.append(temp)
    # for check
    print('\nbest: ')
    pprint(best)
    # make single-string-result: [total-place-num]\n[place1-num]\n[place1-name]\n[place1-address]\n[url] ..
    result_string = str(len(best)) + '\n'
    for i in range(0, len(best)):
        temp1 = str(len(bestLink_list[i])) + '\n' + best[i]['placeName'] + '\n' + best[i]['placeAddress'] + '\n'
        temp2 = ''
        for link in bestLink_list[i]:
            temp2 = temp2 + link + '\n'
        result_string = result_string + temp1 + temp2
        temp1 = ''
    print 'single-string-result: \n' + result_string
    return result_string

# called at app.py
def get_score(_base, _target_list, _APP_ROOT, _user_lat, _user_lon):
    target_list = copy.deepcopy(_target_list)
    for target in target_list:
        target['score'] = how_much_similar(_base, target['jsonDict'])
        target.pop('jsonDict', None)  # remove needless info.
    # sort result with score
    target_list = sorted(target_list, key=lambda d: (d['score']), reverse=True)
    # cut result using cut-score(80% of the max)
    min_index = -1
    for i in range(0, len(target_list)):
        if target_list[i]['score'] < target_list[0]['score'] * 0.8:
            min_index = i
            break
    if min_index > -1:
        target_list[min_index:len(target_list)] = []
    # sort 80%-candidates with distance
    target_list = sort_with_distance(target_list, _user_lat, _user_lon)
    # print and return result
    return print_result(target_list, _APP_ROOT)